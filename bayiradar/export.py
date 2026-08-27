"""Excel / PDF / CSV çıktıları.

PDF'te Türkçe karakter tuzağı: ReportLab'ın gömülü Helvetica fontunda ğ, ş, İ, ı
glifleri yok — kutu olarak basılır. Bu yüzden DejaVuSans'ı elle kaydediyoruz.
"""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .normalize import phone_display

BASLIKLAR = {
    "marka": "Marka", "bayi_adi": "Bayi Adı", "il": "İl", "ilce": "İlçe",
    "adres": "Adres", "telefon": "Telefon", "email": "E-posta",
    "website": "Web Sitesi", "son_gorulme": "Son Güncelleme",
    "veri_durumu": "Veri Durumu",
}
KOLONLAR = list(BASLIKLAR.keys())


def to_dataframe(kayitlar: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(kayitlar)
    for k in KOLONLAR:
        if k not in df.columns:
            df[k] = ""
    df = df[KOLONLAR].rename(columns=BASLIKLAR)
    df["Telefon"] = df["Telefon"].map(phone_display)
    df["Son Güncelleme"] = df["Son Güncelleme"].astype(str).str[:16].str.replace("T", " ")
    return df


# ------------------------------------------------------------------- EXCEL
def to_excel(kayitlar: list[dict], path: str, baslik: str = "") -> str:
    df = to_dataframe(kayitlar)
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="Bayiler", startrow=2)

        # Marka bazlı özet ikinci sekmede
        if not df.empty:
            ozet = (df.groupby("Marka").size().reset_index(name="Bayi Sayısı")
                    .sort_values("Bayi Sayısı", ascending=False))
            ozet.to_excel(xw, index=False, sheet_name="Özet")

        ws = xw.sheets["Bayiler"]
        ws["A1"] = baslik or "Bayi Listesi"
        ws["A1"].font = Font(name="Arial", size=14, bold=True)
        ws["A2"] = f"Oluşturulma: {datetime.now():%d.%m.%Y %H:%M} · {len(df)} kayıt"
        ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")

        hdr_fill = PatternFill("solid", fgColor="1F3864")
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=3, column=c)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = hdr_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        genislik = {"Marka": 18, "Bayi Adı": 38, "İl": 14, "İlçe": 16,
                    "Adres": 55, "Telefon": 17, "E-posta": 28,
                    "Web Sitesi": 28, "Son Güncelleme": 20, "Veri Durumu": 30}
        for i, col in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = genislik.get(col, 18)
            for r in range(4, len(df) + 4):
                ws.cell(row=r, column=i).font = Font(name="Arial", size=10)

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(df.columns))}{len(df) + 3}"

        if "Özet" in xw.sheets:
            wo = xw.sheets["Özet"]
            wo.column_dimensions["A"].width = 24
            wo.column_dimensions["B"].width = 14
            for c in ("A1", "B1"):
                wo[c].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                wo[c].fill = hdr_fill
    return path


# --------------------------------------------------------------------- PDF
_FONT_ADAYLARI = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/Library/Fonts/Arial Unicode.ttf",
    "C:/Windows/Fonts/arial.ttf",
]


def _fontlari_kaydet() -> tuple[str, str]:
    """Türkçe karakterleri destekleyen bir font bulup kaydeder."""
    for p in _FONT_ADAYLARI:
        if Path(p).exists():
            bold = p.replace("DejaVuSans.ttf", "DejaVuSans-Bold.ttf").replace(
                "arial.ttf", "arialbd.ttf")
            pdfmetrics.registerFont(TTFont("TR", p))
            pdfmetrics.registerFont(TTFont("TR-Bold", bold if Path(bold).exists() else p))
            return "TR", "TR-Bold"
    return "Helvetica", "Helvetica-Bold"   # son çare, Türkçe bozuk çıkar


def to_pdf(kayitlar: list[dict], path: str, baslik: str = "") -> str:
    font, font_bold = _fontlari_kaydet()
    df = to_dataframe(kayitlar)
    # PDF'e sığması için dar kolonlar
    gorunen = ["Marka", "Bayi Adı", "İl", "İlçe", "Adres", "Telefon"]
    supheli_satirlar = [i for i, k in enumerate(kayitlar, start=1)
                        if k.get("veri_durumu", "").startswith(
                            ("Son taramada", "Karantina", "Son başarılı", "Eski"))]
    df = df[gorunen]

    doc = SimpleDocTemplate(
        path, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=14 * mm,
        title=baslik or "Bayi Listesi",
    )
    ss = getSampleStyleSheet()
    st_h = ParagraphStyle("h", parent=ss["Title"], fontName=font_bold, fontSize=16)
    st_alt = ParagraphStyle("alt", parent=ss["Normal"], fontName=font,
                            fontSize=8.5, textColor=colors.grey, alignment=TA_CENTER)
    st_hc = ParagraphStyle("hc", fontName=font_bold, fontSize=8.5,
                           textColor=colors.white, leading=10)
    st_c = ParagraphStyle("c", fontName=font, fontSize=8, leading=9.5)

    story = [
        Paragraph(baslik or "Bayi Listesi", st_h),
        Spacer(1, 3),
        Paragraph(f"{datetime.now():%d.%m.%Y %H:%M} · {len(df)} kayıt", st_alt),
        Spacer(1, 8),
    ]

    if df.empty:
        story.append(Paragraph("Bu kriterlere uyan bayi bulunamadı.", st_c))
    else:
        data = [[Paragraph(c, st_hc) for c in df.columns]]
        for _, row in df.iterrows():
            data.append([Paragraph(str(v or "—"), st_c) for v in row])

        t = Table(data, repeatRows=1,
                  colWidths=[28 * mm, 58 * mm, 20 * mm, 24 * mm, 88 * mm, 34 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C8CDD6")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F4F6FA")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ] + [("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFF3D4"))
             for i in supheli_satirlar]
           + [("LINEBEFORE", (0, i), (0, i), 2.5, colors.HexColor("#D89B00"))
             for i in supheli_satirlar]))
        story.append(t)
        if supheli_satirlar:
            story.append(Spacer(1, 8))
            story.append(Paragraph(
                f"Sarı ile işaretli {len(supheli_satirlar)} kayıt son taramada "
                "doğrulanamadı; markanın sitesine ulaşılamadığı için en son "
                "doğrulanmış veri gösteriliyor. Kayıt silinmemiştir.", st_alt))

    def sayfa_alt(canvas, doc_):
        canvas.saveState()
        canvas.setFont(font, 7.5)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(landscape(A4)[0] - 12 * mm, 8 * mm, f"Sayfa {doc_.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=sayfa_alt, onLaterPages=sayfa_alt)
    return path


def to_csv(kayitlar: list[dict], path: str) -> str:
    # Excel'in Türkçe karakterleri doğru açması için BOM'lu UTF-8
    to_dataframe(kayitlar).to_csv(path, index=False, encoding="utf-8-sig", sep=";")
    return path
