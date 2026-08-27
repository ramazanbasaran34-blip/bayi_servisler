"""Tam donanımlı Excel çıktısı.

Üç sayfa:
  Bayi Listesi  – tüm kayıtlar, filtre açık, eski veriler renkli
  Marka Durumu  – hangi marka ne zaman güncellendi (senin sorduğun tazelik bilgisi)
  Nasıl Kullanılır – filtre nasıl kullanılır, renkler ne demek

Makro yok. Makro = Windows'a bağımlılık + her açılışta güvenlik uyarısı +
internetten gelen dosyalarda tam blok. Excel'in kendi filtresi zaten aynı işi
yapıyor ve hiçbir uyarı çıkarmıyor.
"""

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .normalize import phone_display

MUREKKEP = "16202B"
KAGIT = "FCFBF9"
UYARI_ZEMIN = "FDF3E3"
UYARI_YAZI = "A8590C"
CELIK = "5F6E7C"
HAT = "DEDBD4"

KOLONLAR = [
    ("Marka", "marka", 18),
    ("Bayi Adı", "bayi_adi", 38),
    ("İl", "il", 14),
    ("İlçe", "ilce", 16),
    ("Adres", "adres", 52),
    ("Telefon", "telefon", 17),
    ("E-posta", "email", 26),
    ("Web Sitesi", "website", 26),
    ("Veri Durumu", "veri_durumu", 30),
]

ince = Side(style="thin", color=HAT)
KENAR = Border(bottom=ince)


def _baslik_satiri(ws, satir, kolonlar):
    dolgu = PatternFill("solid", fgColor=MUREKKEP)
    for i, ad in enumerate(kolonlar, start=1):
        c = ws.cell(row=satir, column=i, value=ad)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = dolgu
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[satir].height = 22


def to_excel_full(kayitlar, path, marka_durumlari=None,
                  baslik="Motosiklet Yetkili Bayi Rehberi"):
    wb = Workbook()

    # ================================================== 1. BAYİ LİSTESİ
    ws = wb.active
    ws.title = "Bayi Listesi"
    ws.sheet_properties.tabColor = MUREKKEP

    ws["A1"] = baslik
    ws["A1"].font = Font(name="Arial", size=15, bold=True, color=MUREKKEP)
    zaman = datetime.now(timezone.utc).astimezone().strftime("%d.%m.%Y %H:%M")
    supheli_sayi = sum(1 for k in kayitlar
                       if k.get("veri_durumu", "Güncel") != "Güncel")
    ws["A2"] = (f"{len(kayitlar)} bayi · {len({k['marka'] for k in kayitlar})} marka · "
                f"veri {zaman}")
    ws["A2"].font = Font(name="Arial", size=9.5, italic=True, color=CELIK)
    ws["A3"] = ("İl veya İlçe başlığındaki ok işaretine tıklayarak filtreleyin. "
                "Sarı satırlar: son taramada doğrulanamadı, en son bilinen bilgi gösteriliyor.")
    ws["A3"].font = Font(name="Arial", size=9.5, color=UYARI_YAZI if supheli_sayi else CELIK)

    _baslik_satiri(ws, 5, [k[0] for k in KOLONLAR])

    for r, kayit in enumerate(kayitlar, start=6):
        for c, (_, alan, _) in enumerate(KOLONLAR, start=1):
            deger = kayit.get(alan, "")
            if alan == "telefon":
                deger = phone_display(deger)
            hucre = ws.cell(row=r, column=c, value=deger)
            hucre.font = Font(name="Arial", size=10)
            hucre.border = KENAR
            hucre.alignment = Alignment(vertical="top", wrap_text=(alan == "adres"))

    son = len(kayitlar) + 5
    for i, (_, _, gen) in enumerate(KOLONLAR, start=1):
        ws.column_dimensions[get_column_letter(i)].width = gen

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:{get_column_letter(len(KOLONLAR))}{son}"

    # Şüpheli satırları otomatik boya — veri durumu "Güncel" değilse
    if kayitlar:
        ws.conditional_formatting.add(
            f"A6:{get_column_letter(len(KOLONLAR))}{son}",
            FormulaRule(formula=['AND($I6<>"",$I6<>"Güncel")'],
                        fill=PatternFill("solid", fgColor=UYARI_ZEMIN),
                        font=Font(color=UYARI_YAZI)))

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0        # ene sığdır, boyu sayfalara böl
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "5:5"          # her sayfada başlık tekrarlansın

    # ================================================== 2. MARKA DURUMU
    wd = wb.create_sheet("Marka Durumu")
    wd["A1"] = "Marka Bazlı Veri Durumu"
    wd["A1"].font = Font(name="Arial", size=14, bold=True, color=MUREKKEP)
    wd["A2"] = ("Bir markanın sitesi güncellenemediyse kayıtları SİLİNMEZ; "
                "en son doğrulanmış hâliyle listede kalır.")
    wd["A2"].font = Font(name="Arial", size=9.5, italic=True, color=CELIK)

    _baslik_satiri(wd, 4, ["Marka", "Bayi", "Şüpheli", "Son Başarılı Güncelleme",
                           "Son Deneme", "Sonuç", "Not"])
    ETIKET = {"basarili": "Başarılı", "kismi": "Eksik tarama",
              "karantina": "Karantina", "hatali": "Erişilemedi"}
    for r, m in enumerate(marka_durumlari or [], start=5):
        satir = [
            m["marka"], m["toplam"], m["supheli"] or "",
            _tr(m["son_basarili"]), _tr(m["son_deneme"]),
            ETIKET.get(m["son_deneme_durum"], m["son_deneme_durum"] or ""),
            m["son_hata"] or "",
        ]
        sorunlu = m["son_deneme_durum"] not in (None, "", "basarili")
        for c, v in enumerate(satir, start=1):
            h = wd.cell(row=r, column=c, value=v)
            h.font = Font(name="Arial", size=10, bold=(c == 1),
                          color=UYARI_YAZI if sorunlu else "000000")
            h.border = KENAR
            if sorunlu:
                h.fill = PatternFill("solid", fgColor=UYARI_ZEMIN)
    for i, g in enumerate([20, 9, 10, 24, 24, 16, 50], start=1):
        wd.column_dimensions[get_column_letter(i)].width = g
    wd.freeze_panes = "A5"

    # ================================================== 3. REHBER
    wr = wb.create_sheet("Nasıl Kullanılır")
    metin = [
        ("Bu dosya nasıl kullanılır", "h1"),
        ("", ""),
        ("İl veya ilçeye göre liste almak", "h2"),
        ("1. \"Bayi Listesi\" sekmesine geçin.", ""),
        ("2. 5. satırdaki \"İl\" başlığının sağındaki küçük ok işaretine tıklayın.", ""),
        ("3. Açılan listede önce \"Tümünü Seç\"in işaretini kaldırın, "
         "sonra istediğiniz ili işaretleyin.", ""),
        ("4. Aynısını \"İlçe\" için tekrarlayın.", ""),
        ("5. Tablo anında süzülür. Yazdırmak isterseniz Ctrl+P.", ""),
        ("", ""),
        ("Sarı satırlar ne anlama geliyor", "h2"),
        ("O markanın resmi sitesi son güncellemede açılmamış demektir.", ""),
        ("Gösterilen bilgi, en son doğrulanabilen hâlidir; kayıt silinmemiştir.", ""),
        ("Hangi markaların etkilendiğini \"Marka Durumu\" sekmesinde görebilirsiniz.", ""),
        ("", ""),
        ("Veriyi güncellemek", "h2"),
        ("Bu dosyadaki veri, üretildiği andaki veridir.", ""),
        ("Güncellemek için dosyanın bulunduğu klasördeki", ""),
        ("GUNCELLE dosyasına çift tıklayın.", "kalin"),
        ("Siteler yeniden gezilir ve bu dosya yenisiyle değiştirilir.", ""),
        ("İşlem sırasında Excel'i kapatmanız gerekir; dosya üzerine yazılacaktır.", ""),
    ]
    r = 2
    for t, tip in metin:
        h = wr.cell(row=r, column=2, value=t)
        if tip == "h1":
            h.font = Font(name="Arial", size=15, bold=True, color=MUREKKEP)
        elif tip == "h2":
            h.font = Font(name="Arial", size=11.5, bold=True, color=MUREKKEP)
        elif tip == "kalin":
            h.font = Font(name="Arial", size=10.5, bold=True, color=UYARI_YAZI)
        else:
            h.font = Font(name="Arial", size=10.5)
        r += 1
    wr.column_dimensions["A"].width = 3
    wr.column_dimensions["B"].width = 95
    wr.sheet_view.showGridLines = False

    wb.active = 0
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _tr(iso):
    if not iso:
        return "—"
    return datetime.fromisoformat(iso).astimezone().strftime("%d.%m.%Y %H:%M")
